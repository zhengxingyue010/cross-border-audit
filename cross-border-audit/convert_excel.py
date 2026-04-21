#!/usr/bin/env python3
import sys
import csv
from openpyxl import load_workbook

excel_file = "客户订单明细-test.xlsx"
csv_file = "data/raw/客户订单明细.csv"

print(f"正在读取 {excel_file}...")
wb = load_workbook(excel_file)
ws = wb.active

print(f"发现 {ws.max_row} 行, {ws.max_column} 列")

# 提取数据
data = []
for row in ws.iter_rows(values_only=True):
    data.append(row)

# 写入 CSV
print(f"正在写入 {csv_file}...")
with open(csv_file, 'w', newline='', encoding='utf-8-sig') as f:
    writer = csv.writer(f)
    writer.writerows(data)

print("转换完成！")
print(f"\n前 5 行预览：")
for i, row in enumerate(data[:5]):
    print(f"行 {i+1}: {row}")
